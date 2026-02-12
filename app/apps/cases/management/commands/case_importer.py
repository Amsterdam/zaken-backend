from apps.cases.models import Case
from apps.cases.serializers import CaseCreateSerializer
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from rest_framework.test import APIRequestFactory


class Command(BaseCommand):
    help = (
        "Import cases from an Excel file\n\n"
        "Example:\n"
        "  python manage.py case_importer cases.xlsx \\\n"
        "    --username import_user@amsterdam.nl \\\n"
        '    --theme 3 "Kamerverhuur" \\\n'
        '    --reason 32 "Eigen onderzoek" \\\n'
        '    --subject 69 "Doorzon" \\\n'
        '    --project 5 "Project"\\\n'
        "    --dry-run"
    )

    def add_arguments(self, parser):
        parser.add_argument("file", type=str)
        parser.add_argument("--username", type=str, required=True)

        parser.add_argument(
            "--theme",
            nargs=2,
            metavar=("ID", "NAME"),
            required=True,
        )

        parser.add_argument(
            "--reason",
            nargs=2,
            metavar=("ID", "NAME"),
            required=True,
        )

        parser.add_argument(
            "--subject",
            action="append",
            nargs=2,
            metavar=("ID", "NAME"),
            required=True,
        )

        parser.set_defaults(dry_run=True)
        parser.add_argument(
            "--no-dry-run",
            action="store_false",
            dest="dry_run",
        )

        parser.add_argument(
            "--project",
            nargs=2,
            metavar=("ID", "NAME"),
            required=True,
        )

    def handle(self, *args, **options):
        excel_file = options["file"]
        dry_run = options["dry_run"]
        username = options["username"]

        theme_id, theme_name = options["theme"]
        reason_id, reason_name = options["reason"]

        theme_id = int(theme_id)
        reason_id = int(reason_id)

        subjects = [{"id": int(sid), "name": name} for sid, name in options["subject"]]
        subject_ids = [s["id"] for s in subjects]

        project_id, project_name = options["project"]
        project_id = int(project_id)

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            raise CommandError(f"Failed to read Excel file: {e}")

        headers = [cell.value for cell in ws[1]]
        try:
            bag_id_idx = headers.index("bag_id")
            description_idx = headers.index("Omschrijving zaak")
        except ValueError:
            raise CommandError(
                "Excel file must contain 'bag_id' and 'Omschrijving zaak' columns"
            )

        User = get_user_model()
        user = User.objects.get(username=username)

        factory = APIRequestFactory()
        request = factory.post("/")
        request.user = user

        processed = 0
        created = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            bag_id = row[bag_id_idx]
            description = row[description_idx]
            if not bag_id:
                self.stdout.write(self.style.WARNING("Skipping row with empty bag_id"))
                continue

            processed += 1

            data = {
                "bag_id": bag_id,
                "description": description,
                "project": {
                    "id": project_id,
                    "name": project_name,
                },
                "project_id": project_id,
                "reason": {
                    "id": reason_id,
                    "name": reason_name,
                },
                "reason_id": reason_id,
                "subject_ids": subject_ids,
                "subjects": subjects,
                "theme": {
                    "id": theme_id,
                    "name": theme_name,
                },
                "theme_id": theme_id,
            }

            serializer = CaseCreateSerializer(
                data=data,
                context={"request": request},
            )
            serializer.is_valid(raise_exception=True)
            self.stdout.write(f"Creating case for bag_id={bag_id}", ending="")
            case_exists = Case.objects.filter(
                address__bag_id=bag_id,
                description=description,
            ).exists()
            if case_exists:
                self.stdout.write(self.style.WARNING("SKIPPED (case already exists)"))
                continue
            if not dry_run:
                serializer.save()
                created += 1

            self.stdout.write(f"OK bag_id={bag_id}")

        self.stdout.write(
            self.style.SUCCESS(
                f"Done. processed={processed}, created={created}, dry_run={dry_run}"
            )
        )
